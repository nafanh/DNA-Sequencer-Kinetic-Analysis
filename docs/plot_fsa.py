import pandas as pd
import numpy as np
import re
import pprint
import matplotlib.pyplot as plt
import math
import os
from Bio import SeqIO
from collections import defaultdict
from Bio import SeqIO
from collections import defaultdict
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import pprint
from mpl_toolkits import mplot3d
from mpl_toolkits.mplot3d import Axes3D
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import numpy as np

# Appending for multiple substrate multiple product problem
# Have the table with the difference ranges then let the user choose which difference ranges you want to pick
#Edit 2/15/2021 adds the functions from seq_analysis_v1.2 to plot_fsa instead. Imported at the top of seq_analysis_v1.3
#file
def filtered_data(name,time_underscore):
    # file_name = 'Burst on PThio DNA.txt'
    f = open(name, 'r')
    headers = ['Dye', 'Peak Number', 'Height', 'Time', 'Well', 'Area', 'Data Point']
    # data = pd.read_csv('Burst on PThio DNA.txt')
    f.readline()
    col_data = []
    count = 0
    # reads each line in the text file
    for test in f:
        # Splits column data based on header
        # Dye/Sample Peak,Sample File Name,Size, Height, Area,Data Point

        test_set = test.split()
        # Output: ['"B,1"', 'PT_100nM_0_TLD_4.2.19_1_2019-04-02_A05.fsa', '894', '11056', '2524']

        # print(test_set)
        # strips the underscore in long description
        desc_fix = test_set[1].split('_')
        # Output: ['PT', '100nM', '0', 'TLD', '4.2.19', '1', '2019-04-02', 'A05.fsa']

        # print(desc_fix)
        # removes long description name
        test_set.pop(1)
        # Output: ['"B,1"','894', '11056', '2524']
        # print(test_set)
        # print(test_set)

        # inserts time into test_set
        test_set.insert(2, desc_fix[time_underscore])

        # Output:['"B,1"', 'PT_100nM_0_TLD_4.2.19_1_2019-04-02_A05.fsa', '0', '894', '11056', '2524']

        # inserts Well number at the end of test_set
        test_set.insert(3, desc_fix[len(desc_fix) - 1])

        # Output: #['"B,1"', 'PT_100nM_0_TLD_4.2.19_1_2019-04-02_A05.fsa', '0', 'A05.fsa', '894', '11056', '2524']

        # print(test_set)

        # splits dye and peak number and puts them into separate columns
        dye_peak = test_set[0].strip('\"').replace(',', '')
        test_set.pop(0)
        i = 0
        # while i < len(dye_peak):
        #     test_set.insert(i, dye_peak[i])
        #     i += 1
        while i < 1:
            test_set.insert(i, dye_peak[i])
            i += 1

        test_set.insert(1, dye_peak[i:])
        # print(test_set)
        # adding all data to each column
        col_data.append(test_set)
        count += 1
    # pprint.pprint(col_data)
    df = pd.DataFrame(col_data, columns=headers)
    # print(col_data)

    # filters height above user input. Note that if filter height
    # is above internal standard height, then error will raise
    # have to add try/except block here for future use
    return df
    # df_int_all = df.loc[df['Dye'] == 'Y']
    # return df_int_all
    # print()
    # print("These are all the internal std. peaks\n")
    # print('---------------------------------------\n')
    # print(df_int_all)
    # print()
    #min_height = input("Please enter the minimum height (make sure bigger than int std desired): ")
    # while string_alpha_check(min_height):
    #     min_height = input("Not valid integer, please try again: ")

    # min_height = int(min_height)
    # df_hmin = df.loc[df['Height'].astype(int) > min_height]
    # return df_hmin


def int_std_all(df):
    df_int_all = df.loc[df['Dye'] == 'Y']
    return df_int_all

def all_peaks_filtered(df,min_height):
    df_hmin = df.loc[df['Height'].astype(int) > min_height]
    df_hmin_2 = df_hmin.loc[df_hmin['Dye'] == 'B']
    return df_hmin_2

def all_peaks_filtered_global(df,min_height):
    df_hmin = df.loc[df['Height'].astype(int) > min_height]
    return df_hmin
def int_std_filtered(df,min_height_int_std):
    df_int_std_filtered = df.loc[df['Height'].astype(int) > min_height_int_std]
    df_int_std_filtered_2 = df_int_std_filtered.loc[df_int_std_filtered['Dye'] == 'Y']
    return df_int_std_filtered_2


def kintek_export(df):
    col_length = df.shape[1]
    midpt = int((col_length) / 2) + 1
    df1 = df.iloc[:, midpt:]
    # p = Path('Exported Data')
    # p.mkdir(exist_ok=True)

    return df1


filepath = "C:\\Users\\nhuan_000\PycharmProjects\dna seq\pthio\\burst_ranges.xlsx"


def import_ranges(filepath):
    # filepath = "C:\\Users\\nhuan_000\PycharmProjects\dna seq\pthio\\burst_ranges.xlsx"

    wb = load_workbook(filepath)
    ws = wb.active

    size = ws["A"]
    container = []
    lower = ws["B"]
    upper = ws["C"]
    sizes = []
    lower_bounds = []
    upper_bounds = []

    for i in range(1, len(lower)):
        sizes.append(size[i].value)
        lower_bounds.append(lower[i].value)
        upper_bounds.append(upper[i].value)

    container.append(sizes)
    container.append(lower_bounds)
    container.append(upper_bounds)
    # print(lower_bounds)
    # print(upper_bounds)
    new_list = []
    # for i in range(len(lower_bounds)):
    #     name = '[' + str(lower_bounds[i]) + ':' + str(upper_bounds[i]) + ']'
    #     new_list.append(name)

    # pre_list = list(map(list, zip(lower_bounds, upper_bounds)))
    # pre_list = str(pre_list)
    # new_list = pre_list.replace('],', ']:')
    # new_list2 = new_list[1:-1]
    # new_list3 = new_list2.replace(' ', '')
    return container


# print(import_ranges(filepath))
# print(type(import_ranges(filepath)))

def import_sizes(filePath):
    wb = load_workbook(filePath)
    ws = wb.active
    size = ws['A']
    polymer_list = ''
    for i in range(1, len(size)):
        if i == 1:
            polymer_list += str(size[i].value)
        else:
            polymer_list += ',' + str(size[i].value)
    return polymer_list


def atoi(text):
    return int(text) if text.isdigit() else text


def natural_keys(text):
    return [atoi(c) for c in re.split(r'(\d+)', text)]


# if remove_datapt == None:
#     break
# while remove_datapt not in n.columns:
#     remove_datapt = sg.PopupGetText('Polymer not in table, please try again', title='Error')
#     if remove_datapt == None:
#         break
# del n[remove_datapt]
# del n[remove_datapt + '/Total']
# headers_list.remove(remove_datapt)
# # headers_list.remove(remove_datapt + '/Total')
# n['Total'] = 0
# n['Total'] = n.iloc[:, :int(len(n.columns) / 2)].sum(axis=1)
# # remove_datapt_quest= input('Any more polymers you want to delete? Please enter y for yes and n for no: ')
# remove_datapt_quest = sg.PopupGetText(
#     'Any more polymers you want to delete? Please enter y for yes and n for no: ', title='Additional Polymers')


# for i in range(len(headers_list) - 1):
#     divide_name = headers_list[i] + '/' + headers_list[-1]
#     n[divide_name] = n.iloc[:, i] / n.iloc[:, len(headers_list) - 1]
# # Sorts table by time and fill 'NaN' values with '0'
# n = n.fillna(0)
# n.sort_index(inplace=True)
#         n = n.fillna(0)
#         n['Total'] = n.sum(axis=1)
#         headers_list = n.columns.values.tolist()
#         #print(headers_list)
#         for i in range(len(headers_list)-1):
#             divide_name = headers_list[i] + '/' + headers_list[-1]
#             n[divide_name] = n.iloc[:,i] / n.iloc[:,len(headers_list)-1]
#         #n = n.fillna(0)
#         print('This is the table for fractional area vs. size. Please analyze to see if you want to delete outliers.')
#         print('-----------------------------------------------------------------------------')
#         print(n)
#
#         print(headers_list)
#         remove_datapt_quest = input('Do you want to delete a polymer? If so, please enter ''y'', else enter anything else: ')
#         while remove_datapt_quest == 'y' or remove_datapt_quest == 'Y':
#             remove_datapt = input('Please enter the polymer name you want to delete (ex: 24mer): ')
#             while 'mer' not in remove_datapt:
#                 remove_datapt = input('Not valid polymer name, please try again: ')
#             del n[remove_datapt]
#             del n[remove_datapt + '/Total']
#             headers_list.remove(remove_datapt)
#             print(headers_list)
#             print(n)
#             #headers_list.remove(remove_datapt + '/Total')
#             n['Total'] = 0
#             n['Total'] = n.iloc[:,:int(len(n.columns)/2)].sum(axis=1)
#             remove_datapt_quest= input('Any more polymers you want to delete? Please enter y for yes and n for no: ')
#         for i in range(len(headers_list)-1):
#             divide_name = headers_list[i] + '/' + headers_list[-1]
#             n[divide_name] = n.iloc[:,i] / n.iloc[:,len(headers_list)-1]
#         #Sorts table by time and fill 'NaN' values with '0'
#         n = n.fillna(0)
#
#         #Exports non-concentration fixed data. Gives the fractional area for
#         #Each polymer and the time points
#         #export_before_conc = n.to_csv('Export_data_Before_concfix.csv',sep=',')
#         return n
#


def deletePolymer(n, polymers):
    # Test set ['Time','27mer','28mer','Total','27mer/Total','28mer/Total']
    headers_list = n.columns.values.tolist()
    # print(headers_list)
    # print(headers_list) #['Time', '27mer', '28mer', 'Total', '27mer/Total', '28mer/Total']
    headers_list = headers_list[1:(len(headers_list) // 2) + 1]
    # print(headers_list) #['27mer', '28mer', 'Total']
    # print(headers_list) #Output: ['27mer','28mer','Total']
    for remove_datapt in polymers:
        del n[remove_datapt]
        del n[remove_datapt + '/Total']
        headers_list.remove(remove_datapt)
        # print(headers_list) #['28mer', 'Total']
        # Output: ['28mer','Total']
        # headers_list.remove(remove_datapt + '/Total')
        n['Total'] = 0
        n['Total'] = n.iloc[:, 1:int(len(n.columns) / 2) + 1].sum(axis=1)
        # print(n)

    # print(headers_list)
    # print(n)
    for i in range(len(headers_list) - 1):
        divide_name = headers_list[i] + '/' + headers_list[-1]  # last is always total ok
        n[divide_name] = n.loc[:, headers_list[i]] / n.loc[:, 'Total']
    # Sorts table by time and fill 'NaN' values with '0'
    # print(n)
    n = n.fillna(0)
    return n


# n = pd.DataFrame([[0,5,1,2,3,float(5/10),float(1/3),float(2/3)],[0.5,5,6,7,13,float(5/10),float(6/13),float(7/13)]],columns = ['Time','26mer','27mer','28mer','Total','26mer/Total','27mer/Total','28mer/Total'])
# #print(n)
# o = pd.DataFrame([[0,1,2,3,float(1/3),float(2/3)],[0.5,6,7,13,float(6/13),float(7/13)]],columns = ['Time','27mer','28mer','Total','27mer/Total','28mer/Total'])
# p = pd.DataFrame([[0,5,5,1,2,3,float(5/10),float(5/10),float(1/3),float(2/3)],[0.5,5,5,6,7,13,float(5/10),float(5/10),float(6/13),float(7/13)]],columns = ['Time','25mer','26mer','27mer','28mer','Total','25mer/Total','26mer/Total','27mer/Total','28mer/Total'])
#
#
# print(deletePolymer(n,['27mer']))
# print(deletePolymer(o,['27mer']))
# print(deletePolymer(p,['27mer']))

def plot3d(length_dir, fsa_names_sorted, time_underscore, x_min, x_max, y_min, y_max, folder_fsa):
    global x_std_max
    fig2 = plt.figure()
    ax2 = plt.axes(projection='3d')
    # Counter for subplot number
    subplot_num = 1
    # Creates a list for all the differences. Will take the largest difference to set the axes x min and x max (y min and y max for 3d scatter)
    diff_list = []
    # parses the directory for .fsa files
    for k in range(length_dir):
        if k == 0: # This gets the standard reference peak, just the first .fsa file in the directory
            # ax = fig.add_subplot(3,3,subplot_num,projection='3d')
            first_dp_split = fsa_names_sorted[k].split('_')
            time = first_dp_split[time_underscore]
            standard = SeqIO.read(folder_fsa + '/' + fsa_names_sorted[k], 'abi')
            s_abif_key = standard.annotations['abif_raw'].keys()
            s_trace = defaultdict(list)
            s_channels = ['DATA1', 'DATA3']
            for sc in s_channels:
                # s_trace['DATA1'] = y values
                s_trace[sc] = standard.annotations['abif_raw'][sc]
            x_values = np.asarray(s_trace['DATA1'])
            y_std_max = max(s_trace['DATA3'])  # Max y-value for the first graph (used as standard)
            x_std_max = s_trace['DATA3'].index(y_std_max)  # The x-value for the max y-value
            # print(x_values[x_min:x_max+1])

            z_values = x_values[x_min:x_max + 1]  # Makes 3D plot z-values the x-values of 2D plot
            x_values_time = [float(time)] * len(z_values)  # np array with as many time values as there are data points
            y_values = np.arange(x_min, x_max + 1)  # Y-values of 3d plot same as 2d plot
            ##    print(len(x_values))
            ##    print(len(y_values))
            ##    print(len(z_values))
            # x_values = np.arange(1,len(y_values)+1)
            ##
            ##
            sc_std = ax2.plot(x_values_time, y_values, z_values, alpha=0.7)
            continue

        subplot_num += 1
        name_split = fsa_names_sorted[k].split('_')
        time_peak = name_split[time_underscore]
        # opens up the FSA file
        record = SeqIO.read(folder_fsa + '/' + fsa_names_sorted[k], 'abi')
        # Record returns a bunch of dictionaries. Use this line to get the dictionary
        # keys of abif_raw only
        abif_key = record.annotations['abif_raw'].keys()
        # Creates an empty list as the value in the dict
        trace = defaultdict(list)
        # DATA1 is where all the peak value is held, so only grab this dictionary key
        channels = ['DATA1', 'DATA3']
        # Parses the channels list and returns the values for each key in dictionary
        for c in channels:
            trace[c] = record.annotations['abif_raw'][c]
        # Xvalues for time pts
        x_values_non_std = np.asarray(trace['DATA1']) # The actual y-values from 2d plot
        # Numpy for y values (xvalues_non_std)

        # Get the max value data
        y_peak = max(trace['DATA3'])
        # Gets the x value of the max value
        x_peak = trace['DATA3'].index(y_peak)
        # Takes difference of reference x value and time point x value
        diff = x_peak - x_std_max
        diff_list.append(diff)
        # y_values_append_values = np.arange(x_max - diff, x_max)
        #print(diff)

        # X_values_non_std are really the y values on a 2d graph
        y_values_non_std = np.arange(x_min, x_max + 1) - diff # Diff removes the amount of data points +/- from the graph
        # print(y_values_non_std)
        # x_min_diff_first_x = np.where(y_values_non_std == x_min)
        # print(x_min_diff_first_x)
        # if x_min_diff_first_x[0].size != 0:
        #     x_min_diff_first_x_range = np.arange(x_min_diff_first_x[0])
        #     #print(x_min_diff_first_x_range)
        #     y_values_non_std = np.delete(y_values_non_std,x_min_diff_first_x_range)
        # print(y_values_non_std)
        # print(x_min_diff_first_x)
        # print(y_values_non_std)
        # x_min_diff_first_x_np = np.arange(x_min_diff_first_x)
        # np.delete(y_values_non_std,x_min_diff_first_x_np)
        z_values_non_std = x_values_non_std[x_min:x_min + len(y_values_non_std)] # Data points corrected for shift.
        # Adds zero values
        x_values_time_non_std = [float(time_peak)] * len(y_values_non_std) # Makes list of time points with len =
        # data point length

        diff_range_list = np.arange(diff) # makes np array with length of difference, if negative then empty array

        # Removes the values less than x_min for positive differences
        y_values_non_std = np.delete(y_values_non_std, diff_range_list)
        #print(y_values_non_std)
        z_values_non_std = np.delete(z_values_non_std, diff_range_list)
        x_values_time_non_std = np.delete(x_values_time_non_std, diff_range_list)

        # print(diff)

        zero_list = [0] * diff
        # if diff > 0:

        # Positive differences, we will
        if diff > 0:
            y_values_append_values = np.arange(x_max - diff, x_max)
            # z_values_append_values = np.array(zero_list)
            z_values_append_values = np.array(x_values_non_std[x_max:x_max+diff]) # Appends RFU values greater than
            # x_max after shifting
            x_values_time_append_values = np.array([float(time_peak)] * diff)

            y_values_non_std = np.append(y_values_non_std, y_values_append_values)
            z_values_non_std = np.append(z_values_non_std, z_values_append_values)
            x_values_time_non_std = np.append(x_values_time_non_std, x_values_time_append_values)
            # print(len(y_values_non_std))

        else:
            y_values_append_values = np.arange(x_min, x_min - diff + 1) # Subtracting by a negative number is addition
            x_range = x_max - x_min
            y_values_delete_values_end = [x for x in range(x_range + 1, x_range - diff + 2)] # from [end: end+diff]
            # print(y_values_delete_values_end)
            # appended to y_values
            y_values_delete_values_begin = [x for x in range(0, abs(diff) + 1)]
            # y_values_delete_values = np.arange(x_max,x_max - diff)
            z_values_append_values = np.array(zero_list) # Zero values for appending to beginning of z-values (RFU)
            x_values_time_append_values = np.array([float(time_peak)] * diff) # Time values to append)
            # print(x_values_time_append_values)

            y_values_non_std = np.insert(y_values_non_std, 0, y_values_append_values) # insert data point values at
            # beginning of numpy array
            y_values_non_std = np.delete(y_values_non_std, y_values_delete_values_end)
            z_values_non_std = x_values_non_std[x_min:x_min + len(y_values_non_std)]
            x_values_time_non_std = [float(time_peak)] * len(y_values_non_std)


            # print(b)
            # print(len(y_values_non_std))
            z_values_non_std = np.append(z_values_non_std, z_values_append_values) # Add zero values to beginning (
            # negative data point values)
            x_values_time_non_std = np.append(x_values_time_non_std, x_values_time_append_values)
            # x_values_time_non_std = math.log(x_values_time_non_std)*100

        # x_values_time_non_std = np.log(x_values_time_non_std)
        # print(x_values_time_non_std)
        # else:
        #   y_values_append_values = np.arange(x_max + diff, x_max)

        # ax = fig.add_subplot(3,3,subplot_num,projection='3d')
        # np.append(y_values_non_std,zero_list)

        # print(y_values_non_std)

        sc_non_std = ax2.plot(x_values_time_non_std, y_values_non_std, z_values_non_std, alpha=0.7)
    ax2.set_ylim(x_min, x_max)
    ax2.set_zlim(y_min, y_max)
    ax2.set_xlabel('Time')
    ax2.set_ylabel('Data Point')
    ax2.set_zlabel('RFU')
    # make the panes transparent
    # ax.xaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
    # ax.yaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
    # ax.zaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
    # ax.set_facecolor('grey')
    # make the grid lines transparent
    # ax.xaxis._axinfo["grid"]['color'] =  (1,1,1,0)
    # ax.yaxis._axinfo["grid"]['color'] =  (1,1,1,0)
    # ax.zaxis._axinfo["grid"]['color'] =  (1,1,1,0)

    # #Gets x values for vectorization purposes
    # array = np.arange(1,len(trace['DATA1'])+1)
    # #Subtracts difference from array (vectorization)
    # array -= diff

    plt.show()
    #plt.close()


def divisors(n):  # Gives output of a list of lists [[1,4],[2,2]]
    factors = []
    for i in range(1, int(n ** 0.5) + 1):
        if n % i == 0:
            factors.append([i, int(n / i)])

    rev_factors = []
    for i in range(len(factors)):
        rev_factors.append([factors[i][1], factors[i][0]])
    for x in rev_factors:
        if x not in factors: factors.append(x)
    dim_table = pd.DataFrame(factors, columns=['Num of Rows', 'Num of Cols'])
    dim_table.index.name = 'Possibility #'
    dim_table.index += 1
    return dim_table


def getFSANames(dir_name):
    fsa_names = [x for x in os.listdir(dir_name) if x.endswith('.fsa')]
    ex_name = fsa_names[0]
    return ex_name


def getFSANamesListNotNum(dir_name):
    fsa_names = [x for x in os.listdir(dir_name) if x.endswith('.fsa')]
    return fsa_names


def getFSANamesList(dir_name, time_underscore):
    fsa_names = [x for x in os.listdir(dir_name) if x.endswith('.fsa')]
    time_list = []
    # Adjusts to make the zero time point the first .fsa file in the directory
    for i in range(len(fsa_names)):
        name_list = fsa_names[i].split('_')
        time_value = name_list[time_underscore]
        if '.' not in time_value:
            time_list.append(int(time_value))
            continue
        time_list.append(float(time_value))
    # Sorts time_list
    time_list.sort()
    fsa_names_sorted = []

    for time in time_list:
        for ext_name in fsa_names:
            ext_name_split = ext_name.split('_')
            if ext_name_split[time_underscore] == str(time):
                fsa_names_sorted.append(ext_name)
            continue
    # print(fsa_names_sorted)
    return fsa_names_sorted


def timePts(fsa_names, time_underscore):
    time_list = []
    for i in range(len(fsa_names)):
        name_list = fsa_names[i].split('_')
        time_value = name_list[time_underscore]
        if '.' not in time_value:
            time_list.append(int(time_value))
            continue
        time_list.append(float(time_value))
    # Sorts time_list
    time_list.sort()

    return time_list


# Plots 2d alignment
def plot(dir_name, time_underscore, x_min, x_max, y_min, y_max, rows, cols):
    os.chdir(dir_name)
    fsa_names = [x for x in os.listdir(dir_name) if x.endswith('.fsa')]
    length_dir = len(fsa_names)
    time_underscore = int(time_underscore)
    time_list = []
    # Adjusts to make the zero time point the first .fsa file in the directory
    for i in range(len(fsa_names)):
        name_list = fsa_names[i].split('_')
        time_value = name_list[time_underscore]
        if '.' not in time_value:
            time_list.append(int(time_value))
            continue
        time_list.append(float(time_value))
    # Sorts time_list
    time_list.sort()

    fsa_names_sorted = []

    for time in time_list:
        for ext_name in fsa_names:
            ext_name_split = ext_name.split('_')
            if ext_name_split[time_underscore] == str(time):
                fsa_names_sorted.append(ext_name)
            continue

    # Creates figure, axis objects for subplot
    fig, ax = plt.subplots(rows, cols, sharex='all', sharey='all')
    axes_list = [item for sublist in ax for item in sublist]

    # Initialize variables for row and column
    i = 0
    j = 1
    count_1x = 0
    # Parses through each fsa file in the directory
    for k in range(length_dir):
        # n = str(os.getcwd()) + '/' + fsa_dir + '/' + length_dir_name[k]
        # Takes first time point as the standard reference peak
        if k == 0:
            first_dp_split = fsa_names_sorted[k].split('_')
            time = first_dp_split[time_underscore]
            standard = SeqIO.read(fsa_names_sorted[k], 'abi')
            s_abif_key = standard.annotations['abif_raw'].keys()
            s_trace = defaultdict(list)
            s_channels = ['DATA1', 'DATA3']
            for sc in s_channels:
                s_trace[sc] = standard.annotations['abif_raw'][sc]

            # y_val_data = np.asarray(s_trace['DATA3'])
            # print(len(y_val_data))
            y_std_max = max(s_trace['DATA3'])
            x_std_max = s_trace['DATA3'].index(y_std_max)
            # Outputs the graph for the standard peaks
            if rows == 1 or cols == 1:
                ax[count_1x].plot(s_trace['DATA1'], color='black')
                ax[count_1x].set_title('Time: ' + time, loc='right', fontsize=8)
                ax[count_1x].set_xlim(x_min, x_max)
                ax[count_1x].set_ylim(y_min, y_max)
                count_1x += 1

            else:
                ax[0, 0].plot(s_trace['DATA1'], color='black')
                ax[0, 0].set_title('Time: ' + time, loc='right', fontsize=8)
                ax[0, 0].set_xlim(x_min, x_max)
                ax[0, 0].set_ylim(y_min, y_max)
            continue

        # Resets the row once gets to end of row limit
        if j == cols:
            i += 1
            j = 0
        # Gets the time for peaks
        name_split = fsa_names_sorted[k].split('_')
        time_peak = name_split[time_underscore]

        # opens up the FSA file
        record = SeqIO.read(fsa_names_sorted[k], 'abi')
        # Record returns a bunch of dictionaries. Use this line to get the dictionary
        # keys of abif_raw only
        abif_key = record.annotations['abif_raw'].keys()
        # Creates an empty list as the value in the dict
        trace = defaultdict(list)
        # DATA1 is where all the peak value is held, so only grab this dictionary key
        channels = ['DATA1', 'DATA3']
        # Parses the channels list and returns the values for each key in dictionary
        for c in channels:
            trace[c] = record.annotations['abif_raw'][c]

        # Get the max value data
        y_peak = max(trace['DATA3'])
        # Gets the x value of the max value
        x_peak = trace['DATA3'].index(y_peak)
        # Takes difference of reference x value and time point x value
        diff = x_peak - x_std_max
        # print(diff)
        # Gets x values for vectorization purposes
        array = np.arange(1, len(trace['DATA1']) + 1)
        # Subtracts difference from array (vectorization)
        # if diff > 0:
        array -= diff
        # else:
        #     array += diff

        # print(trace['DATA1'])
        # np.append(array,zero_list)
        # Plots the chromatogram data
        ##    for i in range(2):
        ##        for j in range(4):
        ##            if i == 0 and j == 0:
        ##                continue
        if rows == 1:
            ax[count_1x].plot(array, trace['DATA1'], color='black')
            ax[count_1x].set_title('Time: ' + time_peak, loc='right', fontsize=8)
            ax[count_1x].set_xlim(x_min, x_max)
            ax[count_1x].set_ylim(y_min, y_max)
            count_1x += 1
        elif cols == 1:
            ax[count_1x].plot(array, trace['DATA1'], color='black')
            ax[count_1x].set_title('Time: ' + time_peak, loc='right', fontsize=8)
            ax[count_1x].set_xlim(x_min, x_max)
            ax[count_1x].set_ylim(y_min, y_max)
            count_1x += 1
        # Displays the peaks
        else:
            ax[i, j].plot(array, trace['DATA1'], color='black')
            ax[i, j].set_title('Time: ' + time_peak, loc='right', fontsize=8)
            ax[i, j].set_xlim(x_min, x_max)
            ax[i, j].set_ylim(y_min, y_max)
            j += 1
        # Increments column for subplot

    subplot_diff = (rows * cols) - length_dir
    if subplot_diff > 0:
        for i in range(length_dir, (rows * cols)):
            axes_list[i].remove()

    #    plt.plot(array,trace['DATA1'],color='black')
    # plt.xlim(2000,3000)
    #    plt.ylim(0,5000)
    fig.suptitle('Chromatogram Peaks')
    fig.text(0.04, 0.5, 'RFU', va='center', rotation='vertical')

    plt.show()
    #plt.close()


def plot2dScaled(dir_name, time_underscore, x_min, x_max, y_min, y_max, rows, cols, zip_dict, pre_fASB_test):
    os.chdir(dir_name)
    fsa_names = [x for x in os.listdir(dir_name) if x.endswith('.fsa')]
    length_dir = len(fsa_names)
    time_underscore = int(time_underscore)
    time_list = []
    # Adjusts to make the zero time point the first .fsa file in the directory
    for i in range(len(fsa_names)):
        name_list = fsa_names[i].split('_')
        time_value = name_list[time_underscore]
        if '.' not in time_value:
            time_list.append(int(time_value))
            continue
        time_list.append(float(time_value))
    # Sorts time_list
    time_list.sort()

    fsa_names_sorted = []

    for time in time_list:
        for ext_name in fsa_names:
            ext_name_split = ext_name.split('_')
            if ext_name_split[time_underscore] == str(time):
                fsa_names_sorted.append(ext_name)
            continue

    # Creates figure, axis objects for subplot
    fig, ax = plt.subplots(rows, cols, sharex='all', sharey='all')
    axes_list = [item for sublist in ax for item in sublist]

    # Initialize variables for row and column
    i = 0
    j = 1
    count_1x = 0
    # Parses through each fsa file in the directory
    for k in range(length_dir):
        # n = str(os.getcwd()) + '/' + fsa_dir + '/' + length_dir_name[k]
        # Takes first time point as the standard reference peak
        if k == 0:
            first_dp_split = fsa_names_sorted[k].split('_')
            time = first_dp_split[time_underscore]
            standard = SeqIO.read(fsa_names_sorted[k], 'abi')
            s_abif_key = standard.annotations['abif_raw'].keys()
            s_trace = defaultdict(list)
            s_channels = ['DATA1', 'DATA3']
            for sc in s_channels:
                s_trace[sc] = standard.annotations['abif_raw'][sc]

            # y_val_data = np.asarray(s_trace['DATA3'])
            # print(len(y_val_data))
            y_std_max = max(s_trace['DATA3'])
            x_std_max = s_trace['DATA3'].index(y_std_max)
            # Outputs the graph for the standard peaks
            y_std_maxdata1 = max(s_trace['DATA1'])
            x_std_maxdata1 = s_trace['DATA1'].index(y_std_maxdata1)
            size_with_xmax = zip_dict[str(x_std_maxdata1)]  # Gets the size of the xmax value
            size_with_xmax_total = size_with_xmax + '/' + 'Total'
            frac_row = pre_fASB_test.iloc[k, :]
            frac = frac_row.loc[size_with_xmax_total]
            # print(size_with_xmax_total)
            # print(frac)
            # print(y_std_maxdata1)
            factor1 = frac / y_std_maxdata1
            # print(factor1)
            # print(type(s_trace['DATA1']))
            if rows == 1 or cols == 1:
                ax[count_1x].plot(np.array(s_trace['DATA1']) * factor1, color='black')
                ax[count_1x].set_title('Time: ' + time, loc='right', fontsize=8)
                ax[count_1x].set_xlim(x_min, x_max)
                ax[count_1x].set_ylim(y_min, y_max)
                count_1x += 1
            else:
                ax[0, 0].plot(np.array(s_trace['DATA1']) * factor1, color='black')
                ax[0, 0].set_title('Time: ' + time, loc='right', fontsize=8)
                ax[0, 0].set_xlim(x_min, x_max)
                ax[0, 0].set_ylim(y_min, y_max)
            continue

        # Resets the row once gets to end of row limit
        if j == cols:
            i += 1
            j = 0
        # Gets the time for peaks
        name_split = fsa_names_sorted[k].split('_')
        time_peak = name_split[time_underscore]

        # opens up the FSA file
        record = SeqIO.read(fsa_names_sorted[k], 'abi')
        # Record returns a bunch of dictionaries. Use this line to get the dictionary
        # keys of abif_raw only
        abif_key = record.annotations['abif_raw'].keys()
        # Creates an empty list as the value in the dict
        trace = defaultdict(list)
        # DATA1 is where all the peak value is held, so only grab this dictionary key
        channels = ['DATA1', 'DATA3']
        # Parses the channels list and returns the values for each key in dictionary
        for c in channels:
            trace[c] = record.annotations['abif_raw'][c]

        # Get the max value data
        y_peak = max(trace['DATA3'])
        # Gets the x value of the max value
        x_peak = trace['DATA3'].index(y_peak)
        y_peak_data1 = max(trace['DATA1'])  # fractional area/max intenisty
        x_peak_data1 = trace['DATA1'].index(y_peak_data1)

        size_with_xmax_2 = zip_dict[str(x_peak_data1)]  # Gets the size of the xmax value
        size_with_xmax_total_2 = size_with_xmax_2 + '/' + 'Total'
        frac_row_2 = pre_fASB_test.iloc[k, :]
        frac2 = frac_row_2.loc[size_with_xmax_total_2]
        # print(frac2)
        factor2 = frac2 / y_peak_data1
        # print(size_with_xmax_total_2)
        # print(y_peak_data1)
        # print(factor2)
        # Takes difference of reference x value and time point x value
        diff = x_peak - x_std_max
        # print(diff)
        # Gets x values for vectorization purposes
        array = np.arange(1, len(trace['DATA1']) + 1)
        # Subtracts difference from array (vectorization)
        # if diff > 0:
        array -= diff
        # else:
        #     array += diff

        # print(trace['DATA1'])
        # np.append(array,zero_list)
        # Plots the chromatogram data
        ##    for i in range(2):
        ##        for j in range(4):
        ##            if i == 0 and j == 0:
        ##                continue
        if rows == 1:
            ax[count_1x].plot(array, np.array(trace['DATA1']) * factor2, color='black')
            ax[count_1x].set_title('Time: ' + time_peak, loc='right', fontsize=8)
            ax[count_1x].set_xlim(x_min, x_max)
            ax[count_1x].set_ylim(y_min, y_max)
            count_1x += 1
        elif cols == 1:
            ax[count_1x].plot(array, np.array(trace['DATA1']) * factor2, color='black')
            ax[count_1x].set_title('Time: ' + time_peak, loc='right', fontsize=8)
            ax[count_1x].set_xlim(x_min, x_max)
            ax[count_1x].set_ylim(y_min, y_max)
            count_1x += 1
        # Displays the peaks
        else:
            ax[i, j].plot(array, np.array(trace['DATA1']) * factor2, color='black')
            ax[i, j].set_title('Time: ' + time_peak, loc='right', fontsize=8)
            ax[i, j].set_xlim(x_min, x_max)
            ax[i, j].set_ylim(y_min, y_max)
            j += 1
        # Increments column for subplot

    #    plt.plot(array,trace['DATA1'],color='black')
    # plt.xlim(2000,3000)
    #    plt.ylim(0,5000)

    subplot_diff = (rows * cols) - length_dir
    if subplot_diff > 0:
        for i in range(length_dir, (rows * cols)):
            axes_list[i].remove()
    fig.suptitle('Chromatogram Peaks')
    fig.text(0.04, 0.5, 'Frac. Area', va='center', rotation='vertical')

    plt.show()
    #plt.close()


def plot3dScaled(length_dir, fsa_names_sorted, time_underscore, x_min, x_max, y_min, y_max, folder_fsa, zip_dict,
                 pre_fASB_test):  # fvs is a dataframe with the fractional area v size
    fig3 = plt.figure()
    ax3 = plt.axes(projection='3d')
    # Counter for subplot number
    subplot_num = 1
    # Creates a list for all the differences. Will take the largest difference to set the axes x min and x max (y min and y max for 3d scatter)
    diff_list = []
    # parses the directory for .fsa files
    for k in range(length_dir):
        if k == 0:
            # ax = fig.add_subplot(3,3,subplot_num,projection='3d')
            first_dp_split = fsa_names_sorted[k].split('_')
            time = first_dp_split[time_underscore]
            standard = SeqIO.read(folder_fsa + '/' + fsa_names_sorted[k], 'abi')
            s_abif_key = standard.annotations['abif_raw'].keys()
            s_trace = defaultdict(list)
            s_channels = ['DATA1', 'DATA3']
            for sc in s_channels:
                # s_trace['DATA1'] = y values
                s_trace[sc] = standard.annotations['abif_raw'][sc]
            x_values = np.asarray(s_trace['DATA1'])
            y_std_max = max(s_trace['DATA3'])
            x_std_max = s_trace['DATA3'].index(y_std_max)
            # print(x_values[x_min:x_max+1])

            y_std_maxdata1 = max(s_trace['DATA1'])
            x_std_maxdata1 = s_trace['DATA1'].index(y_std_maxdata1)
            size_with_xmax = zip_dict[str(x_std_maxdata1)]  # Gets the size of the xmax value
            size_with_xmax_total = size_with_xmax + '/' + 'Total'
            frac_row = pre_fASB_test.iloc[k, :]
            frac = frac_row.loc[size_with_xmax_total]
            # print(size_with_xmax_total)
            # print(frac)
            # print(y_std_maxdata1)
            factor1 = frac / y_std_maxdata1
            # print(factor1)

            # Gets a range for the max peak

            # Gets max y value for data1:
            y_std_peakdata1 = max(s_trace['DATA1']) / sum(s_trace['DATA1'])
            z_values = x_values[x_min:x_max + 1]
            x_values_time = [float(time)] * len(z_values)
            y_values = np.arange(x_min, x_max + 1)
            ##    print(len(x_values))
            ##    print(len(y_values))
            ##    print(len(z_values))
            # x_values = np.arange(1,len(y_values)+1)
            ##
            ##
            # print(type(z_values))
            scaled_z_values = z_values * (frac / y_std_maxdata1)
            # print(max(z_values))
            # print(max(scaled_z_values))
            # print(scaled_z_values)
            ax3.plot(x_values_time, y_values, scaled_z_values, alpha=0.7)
            continue

        subplot_num += 1
        name_split = fsa_names_sorted[k].split('_')
        time_peak = name_split[time_underscore]
        # opens up the FSA file
        record = SeqIO.read(folder_fsa + '/' + fsa_names_sorted[k], 'abi')
        # Record returns a bunch of dictionaries. Use this line to get the dictionary
        # keys of abif_raw only
        abif_key = record.annotations['abif_raw'].keys()
        # Creates an empty list as the value in the dict
        trace = defaultdict(list)
        # DATA1 is where all the peak value is held, so only grab this dictionary key
        channels = ['DATA1', 'DATA3']
        # Parses the channels list and returns the values for each key in dictionary
        for c in channels:
            trace[c] = record.annotations['abif_raw'][c]
        # Xvalues for time pts
        x_values_non_std = np.asarray(trace['DATA1'])
        # Numpy for y values (xvalues_non_std)

        # Get the max value data
        y_peak = max(trace['DATA3'])
        # Gets the x value of the max value
        x_peak = trace['DATA3'].index(y_peak)

        # Gets the y-peak for Data1
        y_peak_data1 = max(trace['DATA1'])  # fractional area/max intenisty
        x_peak_data1 = trace['DATA1'].index(y_peak_data1)

        size_with_xmax_2 = zip_dict[str(x_peak_data1)]  # Gets the size of the xmax value
        size_with_xmax_total_2 = size_with_xmax_2 + '/' + 'Total'
        frac_row_2 = pre_fASB_test.iloc[k, :]
        frac2 = frac_row_2.loc[size_with_xmax_total_2]
        # print(frac2)
        factor2 = frac2 / y_peak_data1
        # print(size_with_xmax_total_2)
        # print(y_peak_data1)
        # print(factor2)

        # Takes difference of reference x value and time point x value
        diff = x_peak - x_std_max
        diff_list.append(diff)
        # y_values_append_values = np.arange(x_max - diff, x_max)
        # print(diff)

        # X_values_non_std are really the y values on a 2d graph
        y_values_non_std = np.arange(x_min, x_max + 1) - diff

        x_min_diff_first_x = np.where(y_values_non_std == x_min)
        # print(x_min_diff_first_x)
        # if x_min_diff_first_x[0].size != 0:
        #     x_min_diff_first_x_range = np.arange(x_min_diff_first_x[0])
        #     #print(x_min_diff_first_x_range)
        #     y_values_non_std = np.delete(y_values_non_std,x_min_diff_first_x_range)
        # print(y_values_non_std)
        # print(x_min_diff_first_x)
        # print(y_values_non_std)
        # x_min_diff_first_x_np = np.arange(x_min_diff_first_x)
        # np.delete(y_values_non_std,x_min_diff_first_x_np)
        z_values_non_std = x_values_non_std[x_min:x_min + len(y_values_non_std)] # Z-values are the y-values on 2d graph
        x_values_time_non_std = [float(time_peak)] * len(y_values_non_std)

        diff_range_list = np.arange(diff)

        # Removes the values less than x_min for positive differences
        y_values_non_std = np.delete(y_values_non_std, diff_range_list)
        z_values_non_std = np.delete(z_values_non_std, diff_range_list)
        x_values_time_non_std = np.delete(x_values_time_non_std, diff_range_list)

        scaled_z_values_non_std = z_values_non_std * (frac2 / y_peak_data1)
        # print(max(z_values_non_std))
        # print(max(scaled_z_values_non_std))

        # print(diff)

        zero_list = [0] * diff
        # if diff > 0:

        # Positive differences, we will
        if diff > 0:
            y_values_append_values = np.arange(x_max - diff, x_max)
            # z_values_append_values = np.array(zero_list)
            z_values_append_values = np.array(x_values_non_std[x_max:x_max + diff])
            x_values_time_append_values = np.array([float(time_peak)] * diff)

            y_values_non_std = np.append(y_values_non_std, y_values_append_values)
            z_values_non_std = np.append(z_values_non_std, z_values_append_values)
            x_values_time_non_std = np.append(x_values_time_non_std, x_values_time_append_values)
            # print(len(y_values_non_std))

        else:
            y_values_append_values = np.arange(x_min, x_min - diff + 1)
            x_range = x_max - x_min
            y_values_delete_values_end = [x for x in range(x_range + 1, x_range - diff + 2)]
            y_values_delete_values_begin = [x for x in range(0, abs(diff) + 1)]
            # y_values_delete_values = np.arange(x_max,x_max - diff)

            z_values_append_values = np.array(zero_list)
            x_values_time_append_values = np.array([float(time_peak)] * diff)

            y_values_non_std = np.insert(y_values_non_std, 0, y_values_append_values)
            y_values_non_std = np.delete(y_values_non_std, y_values_delete_values_end)
            z_values_non_std = x_values_non_std[x_min:x_min + len(y_values_non_std)]
            x_values_time_non_std = [float(time_peak)] * len(y_values_non_std)
            # print(b)
            # print(len(y_values_non_std))
            z_values_non_std = np.append(z_values_non_std, z_values_append_values)
            x_values_time_non_std = np.append(x_values_time_non_std, x_values_time_append_values)

        # else:
        #   y_values_append_values = np.arange(x_max + diff, x_max)

        # ax = fig.add_subplot(3,3,subplot_num,projection='3d')
        # np.append(y_values_non_std,zero_list)

        # print(y_values_non_std)

        ax3.plot(x_values_time_non_std, y_values_non_std, z_values_non_std * (frac2 / y_peak_data1), alpha=0.7)
    ax3.set_ylim(x_min, x_max)
    ax3.set_zlim(y_min, y_max)
    ax3.set_xlabel('Time')
    ax3.set_ylabel('Data Point')
    ax3.set_zlabel('Frac. Area')
    # make the panes transparent
    # ax.xaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
    # ax.yaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
    # ax.zaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
    # ax.set_facecolor('grey')
    # make the grid lines transparent
    # ax.xaxis._axinfo["grid"]['color'] =  (1,1,1,0)
    # ax.yaxis._axinfo["grid"]['color'] =  (1,1,1,0)
    # ax.zaxis._axinfo["grid"]['color'] =  (1,1,1,0)

    # #Gets x values for vectorization purposes
    # array = np.arange(1,len(trace['DATA1'])+1)
    # #Subtracts difference from array (vectorization)
    # array -= diff

    plt.show()
    #plt.close()


def plot3dlog(length_dir, fsa_names_sorted, time_underscore, x_min, x_max, y_min, y_max, folder_fsa):
    fig2 = plt.figure()
    ax2 = plt.axes(projection='3d')
    # Counter for subplot number
    subplot_num = 1
    # Creates a list for all the differences. Will take the largest difference to set the axes x min and x max (y min and y max for 3d scatter)
    diff_list = []
    # parses the directory for .fsa files
    for k in range(length_dir):
        if k == 0:
            # ax = fig.add_subplot(3,3,subplot_num,projection='3d')
            first_dp_split = fsa_names_sorted[k].split('_')
            time = first_dp_split[time_underscore]
            standard = SeqIO.read(folder_fsa + '/' + fsa_names_sorted[k], 'abi')
            s_abif_key = standard.annotations['abif_raw'].keys()
            s_trace = defaultdict(list)
            s_channels = ['DATA1', 'DATA3']
            for sc in s_channels:
                # s_trace['DATA1'] = y values
                s_trace[sc] = standard.annotations['abif_raw'][sc]
            x_values = np.asarray(s_trace['DATA1'])
            y_std_max = max(s_trace['DATA3'])
            x_std_max = s_trace['DATA3'].index(y_std_max)
            # print(x_values[x_min:x_max+1])

            z_values = x_values[x_min:x_max + 1]
            x_values_time = [float(time)] * len(z_values)
            y_values = np.arange(x_min, x_max + 1)
            ##    print(len(x_values))
            ##    print(len(y_values))
            ##    print(len(z_values))
            # x_values = np.arange(1,len(y_values)+1)
            ##
            ##
            sc_std = ax2.plot(np.log(x_values_time), y_values, z_values, alpha=0.7)
            continue

        subplot_num += 1
        name_split = fsa_names_sorted[k].split('_')
        time_peak = name_split[time_underscore]
        # opens up the FSA file
        record = SeqIO.read(folder_fsa + '/' + fsa_names_sorted[k], 'abi')
        # Record returns a bunch of dictionaries. Use this line to get the dictionary
        # keys of abif_raw only
        abif_key = record.annotations['abif_raw'].keys()
        # Creates an empty list as the value in the dict
        trace = defaultdict(list)
        # DATA1 is where all the peak value is held, so only grab this dictionary key
        channels = ['DATA1', 'DATA3']
        # Parses the channels list and returns the values for each key in dictionary
        for c in channels:
            trace[c] = record.annotations['abif_raw'][c]
        # Xvalues for time pts
        x_values_non_std = np.asarray(trace['DATA1'])
        # Numpy for y values (xvalues_non_std)

        # Get the max value data
        y_peak = max(trace['DATA3'])
        # Gets the x value of the max value
        x_peak = trace['DATA3'].index(y_peak)
        # Takes difference of reference x value and time point x value
        diff = x_peak - x_std_max
        diff_list.append(diff)
        # y_values_append_values = np.arange(x_max - diff, x_max)
        # print(diff)

        # X_values_non_std are really the y values on a 2d graph
        y_values_non_std = np.arange(x_min, x_max + 1) - diff

        x_min_diff_first_x = np.where(y_values_non_std == x_min)
        # print(x_min_diff_first_x)
        # if x_min_diff_first_x[0].size != 0:
        #     x_min_diff_first_x_range = np.arange(x_min_diff_first_x[0])
        #     #print(x_min_diff_first_x_range)
        #     y_values_non_std = np.delete(y_values_non_std,x_min_diff_first_x_range)
        # print(y_values_non_std)
        # print(x_min_diff_first_x)
        # print(y_values_non_std)
        # x_min_diff_first_x_np = np.arange(x_min_diff_first_x)
        # np.delete(y_values_non_std,x_min_diff_first_x_np)
        z_values_non_std = x_values_non_std[x_min:x_min + len(y_values_non_std)]
        x_values_time_non_std = [float(time_peak)] * len(y_values_non_std)

        diff_range_list = np.arange(diff)

        # Removes the values less than x_min for positive differences
        y_values_non_std = np.delete(y_values_non_std, diff_range_list)
        z_values_non_std = np.delete(z_values_non_std, diff_range_list)
        x_values_time_non_std = np.delete(x_values_time_non_std, diff_range_list)

        # print(diff)

        zero_list = [0] * diff
        # if diff > 0:

        # Positive differences, we will
        if diff > 0:
            y_values_append_values = np.arange(x_max - diff, x_max)
            # z_values_append_values = np.array(zero_list)
            z_values_append_values = np.array(x_values_non_std[x_max:x_max + diff])
            x_values_time_append_values = np.array([float(time_peak)] * diff)

            y_values_non_std = np.append(y_values_non_std, y_values_append_values)
            z_values_non_std = np.append(z_values_non_std, z_values_append_values)
            x_values_time_non_std = np.append(x_values_time_non_std, x_values_time_append_values)
            # print(len(y_values_non_std))

        else:
            y_values_append_values = np.arange(x_min, x_min - diff + 1)
            x_range = x_max - x_min
            y_values_delete_values_end = [x for x in range(x_range + 1, x_range - diff + 2)]
            y_values_delete_values_begin = [x for x in range(0, abs(diff) + 1)]
            # y_values_delete_values = np.arange(x_max,x_max - diff)

            z_values_append_values = np.array(zero_list)
            x_values_time_append_values = np.array([float(time_peak)] * diff)

            y_values_non_std = np.insert(y_values_non_std, 0, y_values_append_values)
            y_values_non_std = np.delete(y_values_non_std, y_values_delete_values_end)
            z_values_non_std = x_values_non_std[x_min:x_min + len(y_values_non_std)]
            x_values_time_non_std = [float(time_peak)] * len(y_values_non_std)
            # print(b)
            # print(len(y_values_non_std))
            z_values_non_std = np.append(z_values_non_std, z_values_append_values)
            x_values_time_non_std = np.append(x_values_time_non_std, x_values_time_append_values)
            # x_values_time_non_std = math.log(x_values_time_non_std)*100

        # x_values_time_non_std = np.log(x_values_time_non_std)
        # print(x_values_time_non_std)
        # else:
        #   y_values_append_values = np.arange(x_max + diff, x_max)

        # ax = fig.add_subplot(3,3,subplot_num,projection='3d')
        # np.append(y_values_non_std,zero_list)

        # print(y_values_non_std)

        sc_non_std = ax2.plot(np.log(x_values_time_non_std), y_values_non_std, z_values_non_std, alpha=0.7)
    ax2.set_ylim(x_min, x_max)
    ax2.set_zlim(y_min, y_max)
    ax2.set_xlabel('Log Time')
    ax2.set_ylabel('Data Point')
    ax2.set_zlabel('RFU')
    # make the panes transparent
    # ax.xaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
    # ax.yaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
    # ax.zaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
    # ax.set_facecolor('grey')
    # make the grid lines transparent
    # ax.xaxis._axinfo["grid"]['color'] =  (1,1,1,0)
    # ax.yaxis._axinfo["grid"]['color'] =  (1,1,1,0)
    # ax.zaxis._axinfo["grid"]['color'] =  (1,1,1,0)

    # #Gets x values for vectorization purposes
    # array = np.arange(1,len(trace['DATA1'])+1)
    # #Subtracts difference from array (vectorization)
    # array -= diff

    plt.show()
    #plt.close()


def plot3dlogscaled(length_dir, fsa_names_sorted, time_underscore, x_min, x_max, y_min, y_max, folder_fsa, zip_dict,
                    pre_fASB_test):
    fig3 = plt.figure()
    ax3 = plt.axes(projection='3d')
    # Counter for subplot number
    subplot_num = 1
    # Creates a list for all the differences. Will take the largest difference to set the axes x min and x max (y min and y max for 3d scatter)
    diff_list = []
    # parses the directory for .fsa files
    for k in range(length_dir):
        if k == 0:
            # ax = fig.add_subplot(3,3,subplot_num,projection='3d')
            first_dp_split = fsa_names_sorted[k].split('_')
            time = first_dp_split[time_underscore]
            standard = SeqIO.read(folder_fsa + '/' + fsa_names_sorted[k], 'abi')
            s_abif_key = standard.annotations['abif_raw'].keys()
            s_trace = defaultdict(list)
            s_channels = ['DATA1', 'DATA3']
            for sc in s_channels:
                # s_trace['DATA1'] = y values
                s_trace[sc] = standard.annotations['abif_raw'][sc]
            x_values = np.asarray(s_trace['DATA1'])
            y_std_max = max(s_trace['DATA3'])
            x_std_max = s_trace['DATA3'].index(y_std_max)
            # print(x_values[x_min:x_max+1])

            y_std_maxdata1 = max(s_trace['DATA1'])
            x_std_maxdata1 = s_trace['DATA1'].index(y_std_maxdata1)
            size_with_xmax = zip_dict[str(x_std_maxdata1)]  # Gets the size of the xmax value
            size_with_xmax_total = size_with_xmax + '/' + 'Total'
            frac_row = pre_fASB_test.iloc[k, :]
            frac = frac_row.loc[size_with_xmax_total]
            # print(size_with_xmax_total)
            # print(frac)
            # print(y_std_maxdata1)
            factor1 = frac / y_std_maxdata1
            # print(factor1)

            # Gets a range for the max peak

            # Gets max y value for data1:
            y_std_peakdata1 = max(s_trace['DATA1']) / sum(s_trace['DATA1'])
            z_values = x_values[x_min:x_max + 1]
            x_values_time = [float(time)] * len(z_values)
            y_values = np.arange(x_min, x_max + 1)
            ##    print(len(x_values))
            ##    print(len(y_values))
            ##    print(len(z_values))
            # x_values = np.arange(1,len(y_values)+1)
            ##
            ##
            # print(type(z_values))
            scaled_z_values = z_values * (frac / y_std_maxdata1)
            # print(max(z_values))
            # print(max(scaled_z_values))
            # print(scaled_z_values)
            ax3.plot(np.log(x_values_time), y_values, scaled_z_values, alpha=0.7)
            continue

        subplot_num += 1
        name_split = fsa_names_sorted[k].split('_')
        time_peak = name_split[time_underscore]
        # opens up the FSA file
        record = SeqIO.read(folder_fsa + '/' + fsa_names_sorted[k], 'abi')
        # Record returns a bunch of dictionaries. Use this line to get the dictionary
        # keys of abif_raw only
        abif_key = record.annotations['abif_raw'].keys()
        # Creates an empty list as the value in the dict
        trace = defaultdict(list)
        # DATA1 is where all the peak value is held, so only grab this dictionary key
        channels = ['DATA1', 'DATA3']
        # Parses the channels list and returns the values for each key in dictionary
        for c in channels:
            trace[c] = record.annotations['abif_raw'][c]
        # Xvalues for time pts
        x_values_non_std = np.asarray(trace['DATA1'])
        # Numpy for y values (xvalues_non_std)

        # Get the max value data
        y_peak = max(trace['DATA3'])
        # Gets the x value of the max value
        x_peak = trace['DATA3'].index(y_peak)

        # Gets the y-peak for Data1
        y_peak_data1 = max(trace['DATA1'])  # fractional area/max intenisty
        x_peak_data1 = trace['DATA1'].index(y_peak_data1)

        size_with_xmax_2 = zip_dict[str(x_peak_data1)]  # Gets the size of the xmax value
        size_with_xmax_total_2 = size_with_xmax_2 + '/' + 'Total'
        frac_row_2 = pre_fASB_test.iloc[k, :]
        frac2 = frac_row_2.loc[size_with_xmax_total_2]
        # print(frac2)
        factor2 = frac2 / y_peak_data1
        # print(size_with_xmax_total_2)
        # print(y_peak_data1)
        # print(factor2)

        # Takes difference of reference x value and time point x value
        diff = x_peak - x_std_max
        diff_list.append(diff)
        # y_values_append_values = np.arange(x_max - diff, x_max)
        # print(diff)

        # X_values_non_std are really the y values on a 2d graph
        y_values_non_std = np.arange(x_min, x_max + 1) - diff  # removes values data point values that are +/- the
        # difference. It's subtraction by difference because if positive, then remove data points in front

        x_min_diff_first_x = np.where(y_values_non_std == x_min)
        # print(x_min_diff_first_x)
        # if x_min_diff_first_x[0].size != 0:
        #     x_min_diff_first_x_range = np.arange(x_min_diff_first_x[0])
        #     #print(x_min_diff_first_x_range)
        #     y_values_non_std = np.delete(y_values_non_std,x_min_diff_first_x_range)
        # print(y_values_non_std)
        # print(x_min_diff_first_x)
        # print(y_values_non_std)
        # x_min_diff_first_x_np = np.arange(x_min_diff_first_x)
        # np.delete(y_values_non_std,x_min_diff_first_x_np)
        z_values_non_std = x_values_non_std[x_min:x_min + len(y_values_non_std)]
        x_values_time_non_std = [float(time_peak)] * len(y_values_non_std)

        diff_range_list = np.arange(diff)

        # Removes the values less than x_min for positive differences
        y_values_non_std = np.delete(y_values_non_std, diff_range_list)
        z_values_non_std = np.delete(z_values_non_std, diff_range_list)
        x_values_time_non_std = np.delete(x_values_time_non_std, diff_range_list)

        scaled_z_values_non_std = z_values_non_std * (frac2 / y_peak_data1)
        # print(max(z_values_non_std))
        # print(max(scaled_z_values_non_std))

        # print(diff)

        zero_list = [0] * diff
        # if diff > 0:

        # Positive differences, we will
        if diff > 0:
            y_values_append_values = np.arange(x_max - diff, x_max)
            # z_values_append_values = np.array(zero_list)
            z_values_append_values = np.array(x_values_non_std[x_max:x_max + diff])
            x_values_time_append_values = np.array([float(time_peak)] * diff)

            y_values_non_std = np.append(y_values_non_std, y_values_append_values)
            z_values_non_std = np.append(z_values_non_std, z_values_append_values)
            x_values_time_non_std = np.append(x_values_time_non_std, x_values_time_append_values)
            # print(len(y_values_non_std))

        else:
            y_values_append_values = np.arange(x_min, x_min - diff + 1)
            x_range = x_max - x_min
            y_values_delete_values_end = [x for x in range(x_range + 1, x_range - diff + 2)]
            y_values_delete_values_begin = [x for x in range(0, abs(diff) + 1)]
            # y_values_delete_values = np.arange(x_max,x_max - diff)

            z_values_append_values = np.array(zero_list)
            x_values_time_append_values = np.array([float(time_peak)] * diff)

            y_values_non_std = np.insert(y_values_non_std, 0, y_values_append_values)
            y_values_non_std = np.delete(y_values_non_std, y_values_delete_values_end)
            z_values_non_std = x_values_non_std[x_min:x_min + len(y_values_non_std)]
            x_values_time_non_std = [float(time_peak)] * len(y_values_non_std)
            # print(b)
            # print(len(y_values_non_std))
            z_values_non_std = np.append(z_values_non_std, z_values_append_values)
            x_values_time_non_std = np.append(x_values_time_non_std, x_values_time_append_values)

        # else:
        #   y_values_append_values = np.arange(x_max + diff, x_max)

        # ax = fig.add_subplot(3,3,subplot_num,projection='3d')
        # np.append(y_values_non_std,zero_list)

        # print(y_values_non_std)

        ax3.plot(np.log(x_values_time_non_std), y_values_non_std, z_values_non_std * (frac2 / y_peak_data1), alpha=0.7)
    ax3.set_ylim(x_min, x_max)
    ax3.set_zlim(y_min, y_max)
    ax3.set_xlabel('Log Time')
    ax3.set_ylabel('Data Point')
    ax3.set_zlabel('Frac. Area')
    # make the panes transparent
    # ax.xaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
    # ax.yaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
    # ax.zaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
    # ax.set_facecolor('grey')
    # make the grid lines transparent
    # ax.xaxis._axinfo["grid"]['color'] =  (1,1,1,0)
    # ax.yaxis._axinfo["grid"]['color'] =  (1,1,1,0)
    # ax.zaxis._axinfo["grid"]['color'] =  (1,1,1,0)

    # #Gets x values for vectorization purposes
    # array = np.arange(1,len(trace['DATA1'])+1)
    # #Subtracts difference from array (vectorization)
    # array -= diff

    plt.show()
    #plt.close()

def sample_distance(filtered_data):
    # gets peaks without internal standard
    df_no_int = filtered_data.loc[filtered_data['Dye'] == 'B']
    #print(df_no_int)

    # gets peaks with internal standard
    df_int_std = filtered_data.loc[filtered_data['Dye'] == 'Y']

    #Exports the internal standard data
    #export_int_std = df_int_std.to_csv('Export_data_int_std.csv',sep = ',')
    # print('These are the internal std. peaks after filtering')
    # print('------------------------------------------------')
    # print(df_int_std)
    # print()
    # makes list of int standard data points
    int_stdlist = df_int_std['Data Point'].tolist()

    # makes list of int standard time points
    int_stdtimelist = df_int_std['Time'].tolist()

    # makes list of sample data points
    sample_list = df_no_int['Data Point'].tolist()

    # makes list of sample time points
    sample_timelist = df_no_int['Time'].tolist()
    # print(sample_timelist)

    #zips the internal standard time points and data points into a dictionary
    # Ex: {1:2,3:4}
    int_std_dict = dict(zip(int_stdtimelist, int_stdlist))

    #zips the sample time points and data points into a nested list
    # Ex: [[1,2],[3,4]]
    sample_2d = [list(a) for a in zip(sample_timelist, sample_list)]

    #pprint.pprint(sample_2d)

    # if sample_2d[i][0] in int_std_dict:

    #Creates a column in the pandas dataframe for difference
    #(internal standard - sample) by using datapoint
    #For each time
    diff_list = []
    for i in range(len(sample_2d)):
        #Subtracts the internal std value by the sample datapoint
        #Method is inefficient, could probably use nested list
        # for the internal standards as well
        #diff = int(int_std_dict[sample_2d[i][0]]) - int(sample_2d[i][1])
        diff = round(float(int(int_std_dict[sample_2d[i][0]])/int(sample_2d[i][1])) * 100,2)
        diff_list.append(diff)

    #This is to prevent pandas SettingwithCopyWarning
    df_diff = df_no_int.copy()
    df_diff['Diff'] = diff_list

    # exports data with compared to internal standard
    #export_excel_difference = df_diff.to_csv('Export_data_intstd_Diff.csv',sep=',')

    #Returns a dataframe containing values with differences and
    #No intenal standard column
    return (df_diff)

#Gives table full of table with sizes

def size(df,length,ranges_list):
    # Asks user input for template size
    # original = int(input('Please enter the template size: '))
    # Adjusts the peak number in relation to template size

    # Polymer length list. Ex: [27mer,28mer,29mer]
    #length = []

    # Nested list for the ranges of the difference ranges for each polymer.
    # Ex: For 27mer: [[300,400]]. Where 300 is lower bound and 400 is upper bound
    #ranges_list = []

    # Creates a dictionary for polymer length and difference ranges
    # Ex: {27mer:[300,400],28mer"[450,500]}
    ranges_dict = {}

    # Loop appends each polymer as the key and the range list as the value
    for i in range(len(ranges_list)):
        ranges_dict[length[i]] = ranges_list[i]
    # print(ranges_dict)

    # Takes difference values to list
    diff_list = df['Diff'].astype(float).tolist()

    # Gets the keys of the dictionary which are the polymer sizes
    ranges_keys = list(ranges_dict.keys())
    # print(ranges_keys)
    final_length_list = []
    # print(diff_list)

    # Parses the differences (internal std - sample) and checks to see
    # If within a certain range. If it is, then it is appends the key(size)
    # to final_length_list
    for j in diff_list:
        for i in range(len(ranges_keys)):
            # Gets the value of ranges. Ex: [300,400]
            width = ranges_dict[ranges_keys[i]]
            # print(ranges_dict[ranges_keys[i]])

            # Takes lower bound
            low = width[0]
            # Takes upper bound
            high = width[-1]
            # If the difference is in between these bounds, then
            # append the size to the final_length_list
            if j >= low and j < high:
                final_length_list.append(ranges_keys[i])

    # print(final_length_list)
    # Creates a new column in the dataframe for the size of each sample run
    df["Size"] = final_length_list

    return df



 # Function that creates the table containing area values for each polymer
def table(df):
    #Creates sorted set with time values (unique values only)
    df_time = list(set(df['Time'].tolist()))
    df_time.sort()

    #Creates sorted set with size values (unique values only)
    df_size = list(set(df['Size'].tolist()))
    df_size.sort(key=natural_keys)

    #print(df_size)
    #n = pd.DataFrame(columns = df_size)
    # headings = df_size
    # headings.append('Total')
    # for i in range(len(df_size)-1):
    #     headings.append(str(df_size[i]) + '/' + headings[len(headings)-1-i])
    # n = pd.DataFrame(df_time,columns=['Time'])
    #  for i in range(len(headings)):
    #      n[headings[i]] = np.nan
    #n.set_index('Time',inplace=True)

    area_list = df['Area'].tolist()
    time_list = df['Time'].astype(float).tolist()
    # time_dict = {'time':time_list}

    #Creates dictionary where time and size are keys. Value is area
    size_list = df['Size'].tolist()

    #Creates tuple of time,size pairs
    #This tuple will eventually be the key values for the dictionary
    two_keys = list(zip(time_list,size_list))
    #print(two_keys)

    new_dict = {}

    #Creates new empty dataframe with time values as index
    n = pd.DataFrame()
    n.index.name = 'Time'

    for size in df_size:
        n[size] = ""
    #print(n)
    #Appends area to dictionary containing the time and size keys
    #Ex: {(0.05 sec,27mer),34000, (0.10 sec, 27mer), 20000}
    for i in range(len(two_keys)):
        new_dict[two_keys[i]] = area_list[i]

    #Parses the dictionary and creates a new column in the table
    #for area. Locates the time and size and inputs the area for that
    #those values into the dataframe
    for key in two_keys:
        time = key[0]
        size = key[1]
        n.loc[time,size] = int(new_dict[key])
    #     print(n)
    # print(new_dict)
    # print(len(new_dict))
    # print('This is the table for fractional area vs. size. Please analyze to see if you want to delete outliers.')
    # print('-----------------------------------------------------------------------------')
    # print(n)
    #
    # #Creates new columns containing the area of each polymer/total
    # remove_datapt_quest = input('Do you want to delete a polymer? If so, please enter ''y'', else enter in ''n'': ')
    # while remove_datapt_quest == 'y' or remove_datapt_quest == 'Y':
    #     remove_datapt = input('Please enter the polymer name you want to delete (ex: 24mer): ')
    #     del n[remove_datapt]
    #     remove_datapt_quest= input('Any more polymers you want to delete? Please enter y for yes and n for no: ')
    # #Creates new columns containing the area of each polymer/total
    # n['Total'] = n.sum(axis=1)
    # headers_list = n.columns.values.tolist()
    # for i in range(len(headers_list)-1):
    #     divide_name = headers_list[i] + '/' + headers_list[-1]
    #     n[divide_name] = n.iloc[:,i] / n.iloc[:,len(headers_list)-1]
    #
    #Sorts table by time and fill 'NaN' values with '0'
    n.sort_index(inplace=True)
    n = n.fillna(0)
    n['Total'] = n.sum(axis=1)
    headers_list = n.columns.values.tolist()
    #print(headers_list)
    for i in range(len(headers_list)-1):
        divide_name = headers_list[i] + '/' + headers_list[-1]
        n[divide_name] = round(n.iloc[:,i] / n.iloc[:,len(headers_list)-1],2)
    #n = n.fillna(0)
    # print('This is the table for fractional area vs. size. Please analyze to see if you want to delete outliers.')
    # print('-----------------------------------------------------------------------------')
    # print(n)

    # remove_datapt_quest = input('Do you want to delete a polymer? If so, please enter ''y'', else enter in ''n'': ')
    # while remove_datapt_quest == 'y' or remove_datapt_quest == 'Y':
    #     remove_datapt = input('Please enter the polymer name you want to delete (ex: 24mer): ')
    #     del n[remove_datapt]
    #     del n[remove_datapt + '/Total']
    #     headers_list.remove(remove_datapt)
    #     #headers_list.remove(remove_datapt + '/Total')
    #     n['Total'] = 0
    #     n['Total'] = n.iloc[:,:int(len(n.columns)/2)].sum(axis=1)
    #     remove_datapt_quest= input('Any more polymers you want to delete? Please enter y for yes and n for no: ')
    for i in range(len(headers_list)-1):
        divide_name = headers_list[i] + '/' + headers_list[-1]
        n[divide_name] = round(n.iloc[:,i] / n.iloc[:,len(headers_list)-1],3)
    #Sorts table by time and fill 'NaN' values with '0'
    n = n.fillna(0)

    #Exports non-concentration fixed data. Gives the fractional area for
    #Each polymer and the time points
    #export_before_conc = n.to_csv('Export_data_Before_concfix.csv',sep=',')
    return n



def atoi(text):
    return int(text) if text.isdigit() else text
def natural_keys(text):
    return [atoi(c) for c in re.split(r'(\d+)',text)]



def conc_fix(df,conc):
    #conc = int(input("Please enter the concentration(nM): "))

    #Creates new columns containing polymer/total multiplied by concentration
    headers = df.columns.values.tolist()
    for head in headers:
        if '/Total' in head:
            df.loc[:,head] = df.loc[:,head] * conc
    df = df.fillna(0)
    df.sort_index(inplace=True)

    #csv_name = input("Please enter desired name of exported csv file: ")

    #Gets the time and concentration/total columns only
    col_length = df.shape[1]
    midpt = int((col_length)/2) + 1
    df1 = df.iloc[:,midpt:]
    # p = Path('Exported Data')
    # p.mkdir(exist_ok= True)

    #Exports .txt file of polymer/total
    # export_txt = df1.to_csv(os.getcwd() + '/' + str(p) + '/' + csv_name  +'.txt',sep='\t')
    # #Exports the data to excel sheet
    # export = df.to_csv(os.getcwd() + '/' + str(p) + '/' + csv_name  + '.csv',sep=',')
    return df